# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import CLI_input

SAFE_MODE = False # prevents writing to excel

WB_LOCATION = '/Users/bradpettigrew/Desktop/7-16to7-22automated.xlsx'

wb = openpyxl.load_workbook(WB_LOCATION)
sheet1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])

# TODO standardize how I reference categories, etc. so I reduce refactoring difficulty

# TODO move this to an excel utils module
def coordToCellname(coord):
    """converts tuple coordinate to excel character/number cell ref"""
    colName = get_column_letter(coord[1])
    rowName = str(coord[0])
    return colName + rowName


def generateCellMap(fixedRow, columnTuples, categoryLabels):
    """generates a dict of all row/columns mapped to category"""
    eqnCols = map(lambda letter: column_index_from_string(letter), columnTuples)
    eqnCoords = map(lambda col: (fixedRow, col), eqnCols)
    return dict(zip(categoryLabels, eqnCoords))

# TODO unit tests for this
# expect generateCellMap(args) to equal expected outcome

categories = ['Food', 'Entertainment', 'Transportation', 'Misc']

anchorDict = generateCellMap(9, ('A', 'E', 'I', 'M'), categories)
eqnDict = generateCellMap(4, ('C', 'G', 'K', 'O'), categories)

def findOpenCell(coordinate):
    """takes tuple cell coordinate and returns new tuple coordinate of first empty cell decending from origin"""
    rowIndex = row=coordinate[0]
    while sheet1.cell(row=rowIndex, column=coordinate[1]).value is not None:
        rowIndex += 1
    return (rowIndex, coordinate[1])

def filterCategories(inputDict):
    """returns array of dicts for each category"""
    return map(lambda category: filter(lambda entry: entry['category'] == category, inputDict), categories)

def writeToCategory(array, category):
    """writes all entries of a category to that category column"""
    startCell = anchorDict[category]
    openCell = findOpenCell(startCell)
    idx = 0

    for entry in array:
        name = entry['name']
        price = entry['price']
        date = entry['date']

        sheet1.cell(row=openCell[0]+idx, column=openCell[1]).value = date
        sheet1.cell(row=openCell[0]+idx, column=openCell[1]+1).value = name
        sheet1.cell(row=openCell[0]+idx, column=openCell[1]+2).value = price
        idx += 1
    endCell = (openCell[0]+idx-1, openCell[1]+2)
    updateEqn((startCell[0], startCell[1]+2), endCell, eqnDict[category])


def updateEqn(startCell, endCell, eqnLocation):
    """updates summation equation in excel sheet"""
    newEqn ="=SUM({}:{})".format(coordToCellname(startCell), coordToCellname(endCell))
    sheet1.cell(row=eqnLocation[0], column=eqnLocation[1]).value = newEqn


for array in filterCategories(CLI_input.userInput()):
    if len(array) > 0:
        if not SAFE_MODE:
            writeToCategory(array, array[0]['category'])
        else:
            print array


wb.save(WB_LOCATION)
